"""把抓到的数据渲染成和手工版一致的 xlsx。"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FONT = "Arial"
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="FFEFEFEF")

HEADERS = ["公司", "模型 / API 模型名", "Input", "Cached input", "Output",
           "上下文", "状态 / 下线日期", "备注", "官方页面（核对用）"]


def _money(v):
    """至少两位小数；小额单价按需要扩展到能精确表示为止。"""
    if v is None:
        return "—"
    for dp in (2, 3, 4, 6):
        s = f"{v:,.{dp}f}"
        if abs(float(s.replace(",", "")) - v) < 1e-9:
            return "$" + s
    return f"${v:,.6f}"


def _ctx(v):
    """1048576 这类二进制百万显示成 1M，而不是 1.05M。"""
    if not v:
        return "—"
    v = int(v)
    if v >= 1_000_000:
        binary = round(v / 1048576)
        if binary and abs(v - binary * 1048576) < 1000:
            return f"{binary}M"
        return f"{v / 1_000_000:g}M"
    return f"{v // 1000}K"


def _style(cell, fill, wrap=False, center=False):
    cell.font = Font(name=FONT)
    cell.fill = PatternFill("solid", fgColor="FF" + fill)
    cell.border = BORDER
    cell.alignment = Alignment(
        horizontal="center" if center else "left", vertical="center", wrap_text=wrap)


def build(cfg: dict, retired_cfg: dict, data: dict, date: str, out_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文本模型价格"

    ws["A1"] = (f"TEXT 文本模型 API 价格（USD / 每 1M tokens，标准同步档）"
                f"｜自动更新于 {date}")
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(name=FONT, bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(name=FONT, bold=True)
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    for prov in cfg["providers"]:
        block_start = row
        for m in prov["models"]:
            rec = data.get(m["key"], {})
            fill = prov["color"] if m.get("primary") else prov["color_alt"]
            status = m.get("status", "")

            flag = rec.get("flag_cached")
            cached_txt = flag if flag == "待核" else _money(rec.get("cached"))

            vals = [
                prov["company"],
                rec.get("name", m.get("name", m["key"])),
                _money(rec.get("input")),
                cached_txt,
                _money(rec.get("output")),
                _ctx(rec.get("context")),
                status,
                m.get("note", ""),
                prov["docs"] if m is prov["models"][0] else "",
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                _style(cell, fill, wrap=c in (1, 7, 8, 9), center=c in (3, 4, 5, 6))
            if status.startswith("⚠"):
                ws.cell(row=row, column=7).font = Font(name=FONT, color="FFC00000", bold=True)
            if vals[8]:
                ws.cell(row=row, column=9).hyperlink = vals[8]
                ws.cell(row=row, column=9).font = Font(
                    name=FONT, color="FF0563C1", underline="single")
            row += 1

        if row - 1 > block_start:
            ws.merge_cells(start_row=block_start, start_column=1, end_row=row - 1, end_column=1)
        a = ws.cell(row=block_start, column=1)
        a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        a.font = Font(name=FONT, bold=True)

    for k, v in {"A": 14, "B": 32, "C": 11, "D": 13, "E": 11,
                 "F": 8, "G": 20, "H": 48, "I": 50}.items():
        ws.column_dimensions[k].width = v
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:I{row - 1}"

    # ---- sheet 2 ----
    ws2 = wb.create_sheet("已彻底下线（调用会报错）")
    ws2["A1"] = "以下型号已停止服务，旧代码里如果还写着要改"
    ws2.merge_cells("A1:E1")
    ws2["A1"].font = Font(name=FONT, bold=True, size=12)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(["公司", "已下线模型", "下线日期 / 现状",
                           "官方建议替换为", "官方说明"], 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = Font(name=FONT, bold=True)
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r = 3
    for item in retired_cfg["retired"]:
        for c, v in enumerate(item, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            _style(cell, "FFFFFF", wrap=True)
        if item[4]:
            ws2.cell(row=r, column=5).hyperlink = item[4]
            ws2.cell(row=r, column=5).font = Font(
                name=FONT, color="FF0563C1", underline="single")
        r += 1
    for k, v in {"A": 12, "B": 48, "C": 46, "D": 34, "E": 50}.items():
        ws2.column_dimensions[k].width = v
    ws2.freeze_panes = "A3"

    # ---- sheet 3 ----
    ws3 = wb.create_sheet("数据来源与口径")
    lines = [
        (f"自动更新于 {date}｜主数据源：LiteLLM model_prices_and_context_window.json", True),
        ("", False),
        ("价格为各家标准同步档。Batch 通常再打 5 折（xAI 8 折，DeepSeek 无 Batch 档）；缓存写入另有 1.25x～2x 溢价。", False),
        ("「待核」= 官网未公布该值，脚本不填自动推算值。", False),
        ("「⚠ 日期」= 官方已公告下线，到期后调用报错。", False),
        ("config/models.yaml 里可对单个模型锁定价格，覆盖自动值；每次覆盖都会在变更报告里留痕。", False),
        ("", False),
        ("各家官方页面：", True),
    ]
    r = 1
    for text, bold in lines:
        cell = ws3.cell(row=r, column=1, value=text)
        cell.font = Font(name=FONT, bold=bold)
        r += 1
    for prov in cfg["providers"]:
        cell = ws3.cell(row=r, column=1, value=prov["company"].replace("\n", " "))
        cell.font = Font(name=FONT)
        link = ws3.cell(row=r, column=2, value=prov["docs"])
        link.hyperlink = prov["docs"]
        link.font = Font(name=FONT, color="FF0563C1", underline="single")
        r += 1
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 70

    wb.save(out_path)
